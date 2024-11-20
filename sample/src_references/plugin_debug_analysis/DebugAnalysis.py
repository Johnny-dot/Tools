import json
import pandas as pd
from datetime import datetime, timedelta, timezone
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

import sample.src_references.common.g.G as G

class DebugAnalysis:
    def __init__(self, kb_vo) -> None:
        self.vo = kb_vo
        self.errors = []
        self.kbMgr = G.getG("KBMgr")
        self._uniqueKey = self.vo.getUniqueKey()
        self.logFilePath = None
        self.outputDir = Path(self.vo.getFuncOutPath())

    def calculate_beijing_time(self, timestamp, base_timestamp=946656000000000):
        base_time_utc = datetime.fromtimestamp(base_timestamp / 1_000_000, tz=timezone.utc)
        target_time_utc = base_time_utc + timedelta(microseconds=timestamp)
        beijing_time = target_time_utc + timedelta(hours=8)
        return beijing_time.strftime('%Y-%m-%d %H:%M:%S.%f')

    def analyze_log(self):
        try:
            sourceItems = self.vo.getVal('sourceItems')
            if not sourceItems:
                raise FileNotFoundError("未找到日志文件，请确保已提供日志文件。")

            for log_name, log_file_path in sourceItems.items():
                self.logFilePath = Path(log_file_path)

                if not self.logFilePath.exists():
                    raise FileNotFoundError(f"日志文件未找到：{self.logFilePath}")

                self.outputDir.mkdir(parents=True, exist_ok=True)

                with open(self.logFilePath, 'r', encoding='utf-8') as f:
                    data = json.load(f)

                summary_list = []
                objects_detail_list = []
                object_types = set()

                for scene_value in data['scenes'].values():
                    dialogs = scene_value.get('dialogs', {})
                    for dialog_value in dialogs.values():
                        objects_before = dialog_value.get('objectsBefore', {})
                        object_types.update(objects_before.keys())

                object_types = sorted(object_types)

                for scene_key, scene_value in data['scenes'].items():
                    scene_name = scene_value['scene_name']
                    dialogs = scene_value.get('dialogs', {})
                    if not dialogs:
                        continue
                    for dialog_name, dialog_value in dialogs.items():
                        start_time = self.calculate_beijing_time(dialog_value.get('startTime', 0))
                        end_time = self.calculate_beijing_time(dialog_value.get('endTime', 0))
                        init_duration = dialog_value.get('initDuration', 0) / 1_000_000
                        pre_memory = dialog_value.get('preMemory', 0)
                        post_memory = dialog_value.get('postMemory', 0)
                        memory_change = post_memory - pre_memory
                        draw_calls = dialog_value.get('drawCalls', 0)

                        summary_row = {
                            '场景名称': scene_name,
                            '对话框名称': dialog_name,
                            'InitStart': start_time,
                            'InitEnd': end_time,
                            'Init_Duration(s)': init_duration,
                            'PreMemory（MB）': pre_memory / (1024 * 1024),
                            'PostMemory（MB）': post_memory / (1024 * 1024),
                            'MemoryChange（MB）': memory_change / (1024 * 1024),
                            'DrawCalls': draw_calls,
                            'StartTimeRaw': dialog_value.get('startTime', 0)
                        }
                        summary_list.append(summary_row)

                        detail_row = {
                            '场景名称': scene_name,
                            '对话框名称': dialog_name,
                            'InitStart': start_time
                        }

                        objects_before = dialog_value.get('objectsBefore', {})
                        for obj_type in object_types:
                            count_before = objects_before.get(obj_type, 0)
                            detail_row[obj_type] = count_before

                        detail_row['StartTimeRaw'] = dialog_value.get('startTime', 0)
                        objects_detail_list.append(detail_row)

                summary_list = sorted(summary_list, key=lambda x: x['StartTimeRaw'])
                objects_detail_list = sorted(objects_detail_list, key=lambda x: x['StartTimeRaw'])

                df_summary = pd.DataFrame(summary_list).drop(columns=['StartTimeRaw'])
                df_objects = pd.DataFrame(objects_detail_list).drop(columns=['StartTimeRaw'])

                excel_file_revised = self.outputDir / f"{Path(log_name).stem}_日志分析结果.xlsx"

                with pd.ExcelWriter(excel_file_revised, engine='openpyxl') as writer:
                    df_summary.to_excel(writer, index=False, sheet_name='对话框摘要')
                    df_objects.to_excel(writer, index=False, sheet_name='对象详细信息')

                wb = load_workbook(excel_file_revised)
                ws_summary = wb['对话框摘要']
                ws_objects = wb['对象详细信息']

                self.format_worksheet(ws_summary)
                self.format_worksheet(ws_objects, max_col_width=35)
                self.add_scene_titles(ws_summary)
                self.add_scene_titles(ws_objects)
                wb.save(excel_file_revised)

                G.getG('LogMgr').getLogger(self._uniqueKey).info("日志分析完成，结果已保存到 %s", excel_file_revised)

        except Exception as e:
            G.getG('LogMgr').getLogger(self._uniqueKey).error("日志分析失败：%s", str(e))
            self.errors.append(str(e))

    def format_worksheet(self, ws, max_col_width=35):
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min((max_length + 2), max_col_width)
            ws.column_dimensions[column].width = adjusted_width

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

    def add_scene_titles(self, ws):
        prev_scene_name = None
        scene_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        row_idx = 2
        while row_idx <= ws.max_row:
            scene_name = ws.cell(row=row_idx, column=1).value
            if scene_name != prev_scene_name:
                ws.insert_rows(row_idx)
                scene_cell = ws.cell(row=row_idx, column=1, value=scene_name)
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=ws.max_column)
                scene_cell.font = Font(bold=True)
                scene_cell.alignment = Alignment(horizontal='center', vertical='center')
                scene_cell.fill = scene_fill
                prev_scene_name = scene_name
                row_idx += 1
            row_idx += 1

    def main(self):
        stages = [
            {"msg": '初始化工作路径和配置', 'rate': 0.1},
            {"msg": '开始分析日志数据', 'rate': 0.8},
            {"msg": '任务完成', 'rate': 0.1}
        ]

        self.kbMgr.registerProgress(self._uniqueKey, stages)
        self.kbMgr.onProgressUpdated(self._uniqueKey, 0)
        self.analyze_log()
        self.kbMgr.onProgressUpdated(self._uniqueKey, 1)

        if not self.errors:
            G.getG('LogMgr').getLogger(self._uniqueKey).info("日志分析全部完成，没有发现错误。")
        else:
            G.getG('LogMgr').getLogger(self._uniqueKey).error("日志分析过程中发现错误：%s", self.errors)

        self.kbMgr.onProgressUpdated(self._uniqueKey, 2)
        return self.errors
